import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from customs_pipeline.config import DEFAULT_YEAR, get_input_csv, get_output_xlsx


def csv_to_excel(input_csv=None, output_xlsx=None, year=DEFAULT_YEAR):
    if input_csv is None:
        input_csv = get_input_csv(year)
    if output_xlsx is None:
        output_xlsx = get_output_xlsx(year)

    if not os.path.exists(input_csv):
        print(f"CSV 文件不存在，无法生成 Excel: {input_csv}")
        return False

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}海关公示汇总"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    with open(input_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = cell_align
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"

    wb.save(output_xlsx)
    print(f"已保存: {output_xlsx}")
    return True
